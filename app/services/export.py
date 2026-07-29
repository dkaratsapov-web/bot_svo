"""Выгрузка заявок в XLSX (два листа: Вступление, Консультации)."""

from __future__ import annotations

import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app import texts
from app.db.models import Application, Consultation

MSK = ZoneInfo("Europe/Moscow")

_JOIN_HEADERS = [
    "Тикет", "Дата (МСК)", "ФИО", "Телефон", "Телефон подтверждён",
    "Населённый пункт", "Участник СВО", "Статус", "Повторная", "max_user_id",
]
_CONSULT_HEADERS = [
    "Тикет", "Дата (МСК)", "Направление", "ФИО", "Телефон",
    "Телефон подтверждён", "Статус", "Повторная", "max_user_id",
]


def _fmt_dt(dt: datetime) -> str:
    return dt.astimezone(MSK).strftime("%d.%m.%Y %H:%M")


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    # Автоширина по содержимому (с ограничением)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1]
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    ws.freeze_panes = "A2"


def build_workbook(
    applications: list[Application], consultations: list[Consultation]
) -> bytes:
    wb = Workbook()

    ws_join = wb.active
    ws_join.title = "Вступление"
    _write_sheet(
        ws_join,
        _JOIN_HEADERS,
        [
            [
                a.ticket,
                _fmt_dt(a.created_at),
                a.fio,
                a.phone,
                "да" if a.phone_verified else "нет",
                a.city or "",
                texts.yes_no(a.is_svo_participant),
                a.status.value,
                "да" if a.is_repeat else "нет",
                a.user.max_user_id if a.user else "",
            ]
            for a in applications
        ],
    )

    ws_con = wb.create_sheet("Консультации")
    _write_sheet(
        ws_con,
        _CONSULT_HEADERS,
        [
            [
                c.ticket,
                _fmt_dt(c.created_at),
                texts.DIRECTION_TITLES[c.direction.value],
                c.fio,
                c.phone,
                "да" if c.phone_verified else "нет",
                c.status.value,
                "да" if c.is_repeat else "нет",
                c.user.max_user_id if c.user else "",
            ]
            for c in consultations
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_filename(date_from: datetime | None, date_to: datetime | None) -> str:
    if date_from and date_to:
        return (
            f"applications_{date_from:%Y%m%d}_{date_to:%Y%m%d}.xlsx"
        )
    return f"applications_{datetime.now(MSK):%Y%m%d}.xlsx"
